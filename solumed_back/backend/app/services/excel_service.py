"""
app/services/excel_service.py
==============================
Genera el libro Excel de recepción técnica para presentar ante el INVIMA.
- Una hoja por mes (Ene-2026, Feb-2026, etc.)
- Todos los campos del historial
- Formato profesional con colores por decisión/defecto
- Totales y resumen por hoja
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from collections import defaultdict
from pathlib import Path


# ── Paleta de colores ────────────────────────────────────────
AZUL_HEADER   = "1E3A5F"   # encabezado principal
AZUL_MED      = "2563EB"   # subencabezado
GRIS_CLARO    = "F1F5F9"   # filas pares
BLANCO        = "FFFFFF"
VERDE         = "16A34A"   # Acepta / Vigente
ROJO          = "DC2626"   # Rechaza / Vencido
AMBAR         = "D97706"   # Menor
NARANJA       = "EA580C"   # Mayor
GRIS_TEXTO    = "64748B"

COLORES_DEFECTO = {
    "Ninguno": "16A34A",
    "Menor":   "D97706",
    "Mayor":   "EA580C",
    "Crítico": "DC2626",
}

THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MESES_ES = [
    "Ene","Feb","Mar","Abr","May","Jun",
    "Jul","Ago","Sep","Oct","Nov","Dic"
]

# ── Columnas del informe ─────────────────────────────────────
COLUMNAS = [
    ("N°",                    5),
    ("Fecha Proceso",        14),
    ("Factura",              14),
    ("Proveedor",            22),
    ("Código Producto",      16),
    ("Nombre Producto / Precentacion Comercial",      32),
    ("Concentración",        14),
    ("Forma Farmacéutica",   20),
    ("Lote",                 14),
    ("Vencimiento Prod.",    16),
    ("Cantidad",             10),
    ("N° Muestras",          12),
    ("Temperatura",          12),
    ("Reg. Sanitario",       22),
    ("Estado INVIMA",        16),
    ("Laboratorio",          24),
    ("Principio Activo",     22),
    ("Expediente",           14),
    ("Defecto",              12),
    ("Decisión",             12),
    ("Observaciones",        30),
]

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10) -> Font:
    return Font(name="Arial", bold=bold, color=color, size=size)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _escribir_encabezado_hoja(ws: Worksheet, drogeria: str, mes_label: str):
    """Fila 1: título principal. Fila 2: subtítulo. Fila 3: columnas."""
    num_cols = len(COLUMNAS)
    last_col = get_column_letter(num_cols)

    # Fila 1 — título
    ws.merge_cells(f"A1:{last_col}1")
    c = ws["A1"]
    c.value = f"REGISTRO DE RECEPCIÓN TÉCNICA DE MEDICAMENTOS — {drogeria.upper()}"
    c.font = Font(name="Arial", bold=True, color=BLANCO, size=13)
    c.fill = _fill(AZUL_HEADER)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    # Fila 2 — subtítulo mes
    ws.merge_cells(f"A2:{last_col}2")
    c = ws["A2"]
    c.value = f"Período: {mes_label}  |  Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font = Font(name="Arial", bold=False, color=BLANCO, size=10)
    c.fill = _fill(AZUL_MED)
    c.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Fila 3 — cabeceras de columna
    for col_idx, (nombre, ancho) in enumerate(COLUMNAS, 1):
        cel = ws.cell(row=3, column=col_idx)
        cel.value = nombre
        cel.font = Font(name="Arial", bold=True, color=BLANCO, size=9)
        cel.fill = _fill("334155")
        cel.alignment = _center()
        cel.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"


def _escribir_fila(ws: Worksheet, row: int, n: int, rec: dict):
    fila_par = row % 2 == 0
    bg = GRIS_CLARO if fila_par else BLANCO

    cumple  = rec.get("cumple", "Acepta")
    defecto = rec.get("defectos", "Ninguno")
    estado  = rec.get("estado_invima", "")

    color_cumple  = VERDE if cumple == "Acepta" else ROJO
    color_defecto = COLORES_DEFECTO.get(defecto, GRIS_TEXTO)
    color_invima  = VERDE if "Vigente" in (estado or "") else ROJO

    valores = [
        n,
        rec.get("fecha_proceso", ""),
        rec.get("factura_id", ""),
        rec.get("proveedor", ""),
        rec.get("codigo_producto", ""),
        rec.get("nombre_producto", ""),
        rec.get("concentracion", ""),
        rec.get("forma_farmaceutica", ""),
        rec.get("lote", ""),
        rec.get("vencimiento", ""),
        rec.get("cantidad", ""),
        rec.get("num_muestras", ""),
        rec.get("temperatura", ""),
        rec.get("registro_sanitario", ""),
        estado,
        rec.get("laboratorio", ""),
        rec.get("principio_activo", ""),
        rec.get("expediente", ""),
        defecto,
        cumple,
        rec.get("observaciones", ""),
    ]

    for col_idx, valor in enumerate(valores, 1):
        cel = ws.cell(row=row, column=col_idx)
        cel.value = valor
        cel.border = BORDER
        cel.font = _font(size=9)
        cel.alignment = _left() if col_idx > 2 else _center()
        cel.fill = _fill(bg)

        # Colores semánticos en columnas especiales (Ajustados sin columna Presentación)
        if col_idx == 15:  # Estado INVIMA
            cel.font = Font(name="Arial", bold=True, color=color_invima, size=9)
        elif col_idx == 19:  # Defecto
            cel.font = Font(name="Arial", bold=True, color=color_defecto, size=9)
        elif col_idx == 20:  # Decisión
            cel.font = Font(name="Arial", bold=True, color=color_cumple, size=9)
            cel.alignment = _center()

    ws.row_dimensions[row].height = 18


def _escribir_totales(ws: Worksheet, fila_inicio: int, fila_fin: int, fila_total: int):
    """Fila de totales con fórmulas."""
    num_cols = len(COLUMNAS)
    last_col = get_column_letter(num_cols)
    rango_cumple = f"T{fila_inicio}:T{fila_fin}"

    ws.merge_cells(f"A{fila_total}:S{fila_total}")
    ws[f"A{fila_total}"].value = "RESUMEN DEL MES"
    ws[f"A{fila_total}"].font = Font(name="Arial", bold=True, color=BLANCO, size=10)
    ws[f"A{fila_total}"].fill = _fill(AZUL_HEADER)
    ws[f"A{fila_total}"].alignment = _center()
    ws[f"A{fila_total}"].border = BORDER

    # Totales en columna T (Decisión)
    ws[f"T{fila_total}"].value = f'=COUNTIF(T{fila_inicio}:T{fila_fin},"Acepta")&" ✓ / "&COUNTIF(T{fila_inicio}:T{fila_fin},"Rechaza")&" ✗"'
    ws[f"T{fila_total}"].font = Font(name="Arial", bold=True, color="000000", size=9)
    ws[f"T{fila_total}"].fill = _fill("E2E8F0")
    ws[f"T{fila_total}"].alignment = _center()
    ws[f"T{fila_total}"].border = BORDER

    ws.row_dimensions[fila_total].height = 20


def generar_excel_historial(
    drogeria_nombre: str,
    registros: list[dict],
    ruta_salida: str
) -> str:
    """
    Genera el Excel con una hoja por mes.
    Cada hoja tiene todos los campos + resumen final.
    """
    wb = Workbook()
    wb.remove(wb.active)  # elimina hoja vacía inicial

    # Agrupar por año-mes
    por_mes: dict[str, list[dict]] = defaultdict(list)
    for rec in registros:
        fecha = rec.get("fecha_proceso", "")
        try:
            dt = datetime.strptime(fecha[:10], "%Y-%m-%d")
            clave = f"{dt.year:04d}-{dt.month:02d}"
        except Exception:
            clave = "Sin-Fecha"
        por_mes[clave].append(rec)

    # Ordenar meses cronológicamente
    for clave in sorted(por_mes.keys()):
        registros_mes = por_mes[clave]

        # Nombre de la hoja
        if clave == "Sin-Fecha":
            nombre_hoja = "Sin Fecha"
            mes_label   = "Sin Fecha"
        else:
            anio, mes = clave.split("-")
            mes_label   = f"{MESES_ES[int(mes)-1]}-{anio}"
            nombre_hoja = mes_label

        ws = wb.create_sheet(title=nombre_hoja)
        _escribir_encabezado_hoja(ws, drogeria_nombre, mes_label)

        fila = 4
        fila_inicio = fila
        for n, rec in enumerate(registros_mes, 1):
            _escribir_fila(ws, fila, n, rec)
            fila += 1
        fila_fin = fila - 1

        # Fila de totales
        if fila_inicio <= fila_fin:
            _escribir_totales(ws, fila_inicio, fila_fin, fila + 1)

    # Hoja de resumen general
    ws_resumen = wb.create_sheet(title="Resumen General", index=0)
    _crear_hoja_resumen(ws_resumen, drogeria_nombre, registros, por_mes)

    wb.save(ruta_salida)
    return ruta_salida


def _crear_hoja_resumen(ws: Worksheet, drogeria: str, todos: list[dict], por_mes: dict):
    """Hoja inicial con estadísticas globales."""
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"].value = f"RESUMEN GENERAL — {drogeria.upper()}"
    ws["A1"].font = Font(name="Arial", bold=True, color=BLANCO, size=14)
    ws["A1"].fill = _fill(AZUL_HEADER)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:E2")
    ws["A2"].value = f"Reporte generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", color=BLANCO, size=10)
    ws["A2"].fill = _fill(AZUL_MED)
    ws["A2"].alignment = _center()
    ws.row_dimensions[2].height = 18

    # Totales globales
    total   = len(todos)
    acepta  = sum(1 for r in todos if r.get("cumple") == "Acepta")
    rechaza = total - acepta
    criticos = sum(1 for r in todos if r.get("defectos") == "Crítico")

    stats = [
        ("Total Productos Recibidos", total,    AZUL_MED),
        ("Aceptados",                 acepta,   VERDE),
        ("Rechazados",                rechaza,  ROJO),
        ("Defectos Críticos",         criticos, NARANJA),
    ]

    fila = 4
    for label, valor, color in stats:
        ws.cell(fila, 1).value = label
        ws.cell(fila, 1).font = _font(bold=True, size=11)
        ws.cell(fila, 1).fill = _fill(GRIS_CLARO)
        ws.cell(fila, 1).border = BORDER
        ws.cell(fila, 1).alignment = _left()

        ws.cell(fila, 2).value = valor
        ws.cell(fila, 2).font = Font(name="Arial", bold=True, color=color, size=14)
        ws.cell(fila, 2).fill = _fill(BLANCO)
        ws.cell(fila, 2).border = BORDER
        ws.cell(fila, 2).alignment = _center()
        ws.row_dimensions[fila].height = 22
        fila += 1

    # Tabla por mes
    fila += 1
    headers_resumen = ["Mes", "Total", "Aceptados", "Rechazados", "% Aprobación"]
    for col, h in enumerate(headers_resumen, 1):
        c = ws.cell(fila, col)
        c.value = h
        c.font = Font(name="Arial", bold=True, color=BLANCO, size=10)
        c.fill = _fill("334155")
        c.alignment = _center()
        c.border = BORDER
    ws.row_dimensions[fila].height = 22
    fila += 1

    for clave in sorted(por_mes.keys()):
        recs = por_mes[clave]
        if clave == "Sin-Fecha":
            mes_lbl = "Sin Fecha"
        else:
            anio, mes = clave.split("-")
            mes_lbl = f"{MESES_ES[int(mes)-1]}-{anio}"

        t = len(recs)
        a = sum(1 for r in recs if r.get("cumple") == "Acepta")
        r = t - a
        pct = f"{(a/t*100):.1f}%" if t else "—"
        fila_par = fila % 2 == 0
        bg = GRIS_CLARO if fila_par else BLANCO

        for col, val in enumerate([mes_lbl, t, a, r, pct], 1):
            c = ws.cell(fila, col)
            c.value = val
            c.font = _font(size=10)
            c.fill = _fill(bg)
            c.border = BORDER
            c.alignment = _center()
            if col == 4 and r > 0:
                c.font = Font(name="Arial", bold=True, color=ROJO, size=10)
            if col == 3 and a > 0:
                c.font = Font(name="Arial", bold=True, color=VERDE, size=10)
        ws.row_dimensions[fila].height = 18
        fila += 1
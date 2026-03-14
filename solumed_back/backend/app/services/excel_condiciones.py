"""
app/services/excel_condiciones.py
==================================
Generación del Excel de control ambiental BPA con formato visual de grilla.

El archivo replica fielmente la planilla física colombiana:
  - Sección TEMPERATURA (eje X: 15–35°C, columnas angostas)
  - Sección HUMEDAD     (eje X: 35–75%, saltos de 5)
  - Cada día tiene 2 filas: M (Mañana / AM) y T (Tarde / PM)
  - La celda correspondiente al valor medido se rellena con color:
      AM → Azul marino #1F3864
      PM → Rojo        #C00000
  - Límites BPA marcados con borde lateral más grueso
  - Valores fuera de rango BPA → borde rojo alrededor de la celda

Uso:
    from app.services.excel_condiciones import generar_excel_control_ambiental
    buf = generar_excel_control_ambiental(registros, mes, anio, nombre, responsable)
    return StreamingResponse(buf, media_type="...xlsx")
"""

import calendar
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Constantes de layout ──────────────────────────────────────────────────

# Valores del eje horizontal — TEMPERATURA: 15, 16, …, 35 (21 columnas)
TEMP_VALS: list[int] = list(range(15, 36))
# Valores del eje horizontal — HUMEDAD: 35, 40, …, 75  (9 columnas)
HUM_VALS:  list[int] = [35, 40, 45, 50, 55, 60, 65, 70, 75]

# Columnas en la hoja (1-based)
C_T_DIA  = 1                          # DIA  — sección temperatura
C_T_MT   = 2                          # M/T  — sección temperatura
C_T_INI  = 3                          # primer valor temp (15°C)
C_T_FIN  = C_T_INI + len(TEMP_VALS) - 1   # último  (35°C) = 23
C_SEP    = C_T_FIN + 1                # columna separadora = 24
C_H_DIA  = C_SEP + 1                  # DIA  — sección humedad = 25
C_H_MT   = C_SEP + 2                  # M/T  — sección humedad = 26
C_H_INI  = C_SEP + 3                  # primer valor hum (35%) = 27
C_H_FIN  = C_H_INI + len(HUM_VALS) - 1    # último  (75%)    = 35
C_TOTAL  = C_H_FIN                    # 35 columnas en total

# Filas fijas de encabezado
F_TITULO  = 1   # Nombre droguería
F_INFO    = 2   # Responsable / Año / Mes
F_SECC    = 3   # "TEMPERATURA" | "HUMEDAD"
F_NUMS    = 4   # 15,16,...,35 | 35,40,...,75  (rotados 90°)
F_INI     = 5   # primera fila de datos

# Colores BPA
HEX_AM     = "1F3864"   # Azul marino oscuro — turno Mañana
HEX_PM     = "C00000"   # Rojo oscuro         — turno Tarde
HEX_GRIS   = "F2F2F2"   # Gris claro          — fondo encabezado números
HEX_BLANCO = "FFFFFF"

# Rangos BPA
BPA_T_MIN = 15
BPA_T_MAX = 30
BPA_H_MAX = 75

MESES_ES = {
    1: "ENERO",    2: "FEBRERO", 3: "MARZO",     4: "ABRIL",
    5: "MAYO",     6: "JUNIO",   7: "JULIO",     8: "AGOSTO",
    9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}


# ── Helpers de estilo ─────────────────────────────────────────────────────

def _side(style: str = "thin", color: str = "000000") -> Side:
    return Side(style=style, color=color)

def _borde(izq="thin", der="thin", arr="thin", aba="thin") -> Border:
    return Border(
        left=_side(izq), right=_side(der),
        top=_side(arr),  bottom=_side(aba),
    )

BORDE_FINO   = _borde()
BORDE_ALERTA = Border(
    left=_side("medium", "FF0000"), right=_side("medium", "FF0000"),
    top=_side("medium", "FF0000"),  bottom=_side("medium", "FF0000"),
)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

FILL_AM     = _fill(HEX_AM)
FILL_PM     = _fill(HEX_PM)
FILL_GRIS   = _fill(HEX_GRIS)
FILL_BLANCO = _fill(HEX_BLANCO)

ALIN_CENTRO = Alignment(horizontal="center", vertical="center")
ALIN_ROT90  = Alignment(horizontal="center", vertical="bottom",
                        text_rotation=90, wrap_text=False)
ALIN_IZQ    = Alignment(horizontal="left", vertical="center")


# ── Cálculo de columna destino ────────────────────────────────────────────

def _col_temp(valor: float) -> int | None:
    """Columna (1-based) para un valor de temperatura. None si está fuera de rango."""
    v = round(valor)
    if v not in TEMP_VALS:
        return None
    return C_T_INI + TEMP_VALS.index(v)


def _col_hum(valor: float) -> int | None:
    """Columna (1-based) para un valor de humedad. Redondea al múltiplo de 5 más cercano."""
    v = round(round(valor / 5) * 5)
    if v not in HUM_VALS:
        return None
    return C_H_INI + HUM_VALS.index(v)


def _borde_temp(val: int) -> Border:
    """Borde de una celda de temperatura, con laterales BPA más gruesos."""
    izq = "medium" if val == BPA_T_MIN else "thin"
    der = "medium" if val == BPA_T_MAX else "thin"
    return _borde(izq, der)


def _borde_hum(val: int) -> Border:
    """Borde de una celda de humedad, con lateral derecho BPA más grueso en 75%."""
    der = "medium" if val == BPA_H_MAX else "thin"
    return _borde(der=der)


# ── Función principal ─────────────────────────────────────────────────────

def generar_excel_control_ambiental(
    registros: list[dict],
    mes: int,
    anio: int,
    nombre_drogueria: str,
    responsable: str = "",
) -> io.BytesIO:
    """
    Genera el archivo Excel del control ambiental BPA con formato visual de grilla.

    Parámetros:
        registros        : lista de dicts → {fecha, temperatura_am, temperatura_pm,
                           humedad_am, humedad_pm}. Las fechas deben ser YYYY-MM-DD.
        mes              : número de mes (1-12)
        anio             : año (ej: 2026)
        nombre_drogueria : nombre que aparece en el encabezado
        responsable      : nombre del responsable del período (opcional)

    Retorna:
        BytesIO con el archivo .xlsx en memoria, listo para StreamingResponse.
    """
    nombre_mes = MESES_ES.get(mes, str(mes))
    dias_mes   = calendar.monthrange(anio, mes)[1]

    wb = Workbook()
    ws = wb.active
    ws.title = f"Control {nombre_mes[:3]} {anio}"

    # ── 1. Anchos de columnas ────────────────────────────────────
    ws.column_dimensions[get_column_letter(C_T_DIA)].width = 4.2
    ws.column_dimensions[get_column_letter(C_T_MT)].width  = 2.5
    for i in range(len(TEMP_VALS)):
        ws.column_dimensions[get_column_letter(C_T_INI + i)].width = 2.5
    ws.column_dimensions[get_column_letter(C_SEP)].width   = 3
    ws.column_dimensions[get_column_letter(C_H_DIA)].width = 4.2
    ws.column_dimensions[get_column_letter(C_H_MT)].width  = 2.5
    for i in range(len(HUM_VALS)):
        ws.column_dimensions[get_column_letter(C_H_INI + i)].width = 2.5

    # ── 2. Fila 1 — Título: nombre de la droguería ──────────────
    ws.row_dimensions[F_TITULO].height = 22
    ws.merge_cells(start_row=F_TITULO, start_column=1,
                   end_row=F_TITULO,   end_column=C_TOTAL)
    c = ws.cell(F_TITULO, 1, f"DROGUERÍA: {nombre_drogueria.upper()}")
    c.font      = Font(bold=True, size=12)
    c.alignment = ALIN_IZQ
    c.fill      = FILL_BLANCO

    # ── 3. Fila 2 — Info: responsable / año / mes ───────────────
    ws.row_dimensions[F_INFO].height = 16
    # Lado izquierdo (sección temperatura)
    ws.merge_cells(start_row=F_INFO, start_column=1,
                   end_row=F_INFO,   end_column=C_T_FIN)
    c = ws.cell(F_INFO, 1, f"Responsable: {responsable}")
    c.font      = Font(size=9)
    c.alignment = ALIN_IZQ
    c.fill      = FILL_BLANCO
    # Lado derecho (sección humedad)
    ws.merge_cells(start_row=F_INFO, start_column=C_H_DIA,
                   end_row=F_INFO,   end_column=C_TOTAL)
    c = ws.cell(F_INFO, C_H_DIA, f"Año: {anio}    Mes: {nombre_mes}")
    c.font      = Font(size=9)
    c.alignment = ALIN_IZQ
    c.fill      = FILL_BLANCO
    # Borde inferior grueso para separar encabezado
    borde_sep = Border(bottom=_side("medium"))
    for col in range(1, C_TOTAL + 1):
        ws.cell(F_INFO, col).border = borde_sep

    # ── 4. Fila 3 — Títulos de sección ──────────────────────────
    ws.row_dimensions[F_SECC].height = 16
    # TEMPERATURA
    ws.merge_cells(start_row=F_SECC, start_column=C_T_DIA,
                   end_row=F_SECC,   end_column=C_T_FIN)
    c = ws.cell(F_SECC, C_T_DIA, "TEMPERATURA (°C)")
    c.font      = Font(bold=True, size=9)
    c.alignment = ALIN_CENTRO
    c.fill      = FILL_BLANCO
    c.border    = Border(bottom=_side("thin"))
    # HUMEDAD
    ws.merge_cells(start_row=F_SECC, start_column=C_H_DIA,
                   end_row=F_SECC,   end_column=C_H_FIN)
    c = ws.cell(F_SECC, C_H_DIA, "HUMEDAD RELATIVA (%)")
    c.font      = Font(bold=True, size=9)
    c.alignment = ALIN_CENTRO
    c.fill      = FILL_BLANCO
    c.border    = Border(bottom=_side("thin"))

    # ── 5. Fila 4 — Encabezado de valores (números rotados 90°) ─
    ws.row_dimensions[F_NUMS].height = 44

    font_num  = Font(bold=True, size=7)

    # Cabecera "DIA" y "" de temperatura
    c = ws.cell(F_NUMS, C_T_DIA, "DIA")
    c.font = Font(bold=True, size=7); c.alignment = ALIN_CENTRO
    c.fill = FILL_GRIS; c.border = BORDE_FINO

    c = ws.cell(F_NUMS, C_T_MT, "")
    c.fill = FILL_GRIS; c.border = BORDE_FINO

    # Números de temperatura con indicadores BPA
    for i, val in enumerate(TEMP_VALS):
        col = C_T_INI + i
        c   = ws.cell(F_NUMS, col, val)
        c.font      = font_num
        c.alignment = ALIN_ROT90
        c.fill      = FILL_GRIS
        c.border    = _borde_temp(val)

    # Cabecera "DIA" y "" de humedad
    c = ws.cell(F_NUMS, C_H_DIA, "DIA")
    c.font = Font(bold=True, size=7); c.alignment = ALIN_CENTRO
    c.fill = FILL_GRIS; c.border = BORDE_FINO

    c = ws.cell(F_NUMS, C_H_MT, "")
    c.fill = FILL_GRIS; c.border = BORDE_FINO

    # Números de humedad con indicador BPA
    for i, val in enumerate(HUM_VALS):
        col = C_H_INI + i
        c   = ws.cell(F_NUMS, col, val)
        c.font      = font_num
        c.alignment = ALIN_ROT90
        c.fill      = FILL_GRIS
        c.border    = _borde_hum(val)

    # ── 6. Construir índice de datos BD ─────────────────────────
    # clave: "YYYY-MM-DD" → dict del registro
    datos: dict[str, dict] = {}
    for r in registros:
        fecha = str(r.get("fecha", "")).strip()
        if fecha:
            datos[fecha] = r

    # ── 7. Dibujar filas de días ─────────────────────────────────
    font_dia = Font(bold=True, size=8)
    font_mt  = Font(size=7)

    for dia in range(1, dias_mes + 1):
        fecha_str = f"{anio}-{mes:02d}-{dia:02d}"
        reg       = datos.get(fecha_str, {})

        fila_m = F_INI + (dia - 1) * 2   # fila Mañana (AM)
        fila_t = fila_m + 1              # fila Tarde   (PM)

        ws.row_dimensions[fila_m].height = 12
        ws.row_dimensions[fila_t].height = 12

        # ── 7a. Columnas DIA (merge 2 filas) ────────────────────
        for c_dia in (C_T_DIA, C_H_DIA):
            ws.merge_cells(start_row=fila_m, start_column=c_dia,
                           end_row=fila_t,   end_column=c_dia)
            c = ws.cell(fila_m, c_dia, dia)
            c.font      = font_dia
            c.alignment = ALIN_CENTRO
            c.fill      = FILL_BLANCO
            c.border    = BORDE_FINO

        # ── 7b. Columnas M/T ────────────────────────────────────
        for fila, letra in ((fila_m, "M"), (fila_t, "T")):
            for c_mt in (C_T_MT, C_H_MT):
                c = ws.cell(fila, c_mt, letra)
                c.font      = font_mt
                c.alignment = ALIN_CENTRO
                c.fill      = FILL_BLANCO
                c.border    = BORDE_FINO

        # ── 7c. Grilla TEMPERATURA ───────────────────────────────
        temp_am = reg.get("temperatura_am")
        temp_pm = reg.get("temperatura_pm")

        # Columna que debe marcarse para cada turno
        col_am_t = _col_temp(temp_am) if temp_am is not None else None
        col_pm_t = _col_temp(temp_pm) if temp_pm is not None else None

        # ¿Está fuera del rango BPA?
        alerta_am_t = temp_am is not None and (temp_am < BPA_T_MIN or temp_am > BPA_T_MAX)
        alerta_pm_t = temp_pm is not None and (temp_pm < BPA_T_MIN or temp_pm > BPA_T_MAX)

        for i, val in enumerate(TEMP_VALS):
            col     = C_T_INI + i
            borde_v = _borde_temp(val)  # borde estándar con marcas BPA

            # — Fila M —
            c = ws.cell(fila_m, col, "")
            if col == col_am_t:
                c.fill   = FILL_AM
                c.border = BORDE_ALERTA if alerta_am_t else borde_v
            else:
                c.fill   = FILL_BLANCO
                c.border = borde_v

            # — Fila T —
            c = ws.cell(fila_t, col, "")
            if col == col_pm_t:
                c.fill   = FILL_PM
                c.border = BORDE_ALERTA if alerta_pm_t else borde_v
            else:
                c.fill   = FILL_BLANCO
                c.border = borde_v

        # ── 7d. Grilla HUMEDAD ───────────────────────────────────
        hum_am = reg.get("humedad_am")
        hum_pm = reg.get("humedad_pm")

        col_am_h = _col_hum(hum_am) if hum_am is not None else None
        col_pm_h = _col_hum(hum_pm) if hum_pm is not None else None

        alerta_am_h = hum_am is not None and hum_am > BPA_H_MAX
        alerta_pm_h = hum_pm is not None and hum_pm > BPA_H_MAX

        for i, val in enumerate(HUM_VALS):
            col     = C_H_INI + i
            borde_v = _borde_hum(val)

            # — Fila M —
            c = ws.cell(fila_m, col, "")
            if col == col_am_h:
                c.fill   = FILL_AM
                c.border = BORDE_ALERTA if alerta_am_h else borde_v
            else:
                c.fill   = FILL_BLANCO
                c.border = borde_v

            # — Fila T —
            c = ws.cell(fila_t, col, "")
            if col == col_pm_h:
                c.fill   = FILL_PM
                c.border = BORDE_ALERTA if alerta_pm_h else borde_v
            else:
                c.fill   = FILL_BLANCO
                c.border = borde_v

    # ── 8. Leyenda al pie ────────────────────────────────────────
    fila_ley = F_INI + dias_mes * 2 + 1
    ws.row_dimensions[fila_ley].height     = 13
    ws.row_dimensions[fila_ley + 1].height = 13

    for fila_l, fill_l, texto in (
        (fila_ley,     FILL_AM, "  Turno Mañana (AM)"),
        (fila_ley + 1, FILL_PM, "  Turno Tarde  (PM)"),
    ):
        # Cuadro de color
        c = ws.cell(fila_l, C_T_DIA, "")
        c.fill   = fill_l
        c.border = BORDE_FINO
        # Texto explicativo (merge para que no tape bordes de grilla)
        ws.merge_cells(start_row=fila_l, start_column=C_T_MT,
                       end_row=fila_l,   end_column=C_T_MT + 5)
        c = ws.cell(fila_l, C_T_MT, texto)
        c.font      = Font(size=7, italic=True)
        c.alignment = Alignment(vertical="center")

    # Nota BPA
    fila_nota = fila_ley + 3
    ws.row_dimensions[fila_nota].height = 12
    ws.merge_cells(start_row=fila_nota, start_column=1,
                   end_row=fila_nota,   end_column=C_T_FIN)
    c = ws.cell(fila_nota, 1,
                "Rango BPA: Temperatura 15–30°C  |  Humedad máx. 75%  "
                "|  Borde rojo = fuera de rango")
    c.font      = Font(size=7, italic=True, color="808080")
    c.alignment = ALIN_IZQ

    # ── 9. Guardar en BytesIO y retornar ─────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

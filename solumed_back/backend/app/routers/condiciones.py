"""
app/routers/condiciones.py
==========================
Rutas para la gestión de condiciones ambientales (temperatura y humedad).
"""
import calendar
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
import io

from app.core.auth import get_usuario_actual
from app.core.database import (
    guardar_condiciones_dia,
    obtener_condiciones_mes,
    verificar_alerta_condiciones,
    get_drogeria,
)
from app.models.schemas import CondicionAmbientalCreate

router = APIRouter()

@router.get("")
def listar_mes(
    mes: str = Query(..., description="Mes en formato YYYY-MM"),
    u: dict = Depends(get_usuario_actual)
):
    """Obtiene todos los registros de condiciones de un mes específico."""
    drogeria_id = u.get("drogeria_id")
    if not drogeria_id:
        raise HTTPException(400, "Superadmin no puede consultar esto.")
    
    condiciones = obtener_condiciones_mes(drogeria_id, mes)
    return {"ok": True, "datos": condiciones}


@router.post("")
def guardar_dia(
    body: CondicionAmbientalCreate,
    u: dict = Depends(get_usuario_actual)
):
    """Guarda o actualiza las condiciones para un día específico."""
    drogeria_id = u.get("drogeria_id")
    if not drogeria_id:
        raise HTTPException(400, "Superadmin no puede guardar esto.")

    try:
        guardar_condiciones_dia(
            drogeria_id=drogeria_id,
            usuario_id=u["id"],
            fecha=body.fecha,
            temperatura_am=body.temperatura_am,
            temperatura_pm=body.temperatura_pm,
            humedad_am=body.humedad_am,
            humedad_pm=body.humedad_pm,
            firma_am=body.firma_am,
            firma_pm=body.firma_pm
        )
        return {"ok": True, "mensaje": "Condiciones guardadas correctamente"}
    except Exception as e:
        raise HTTPException(500, f"Error guardando condiciones: {e}")


@router.get("/alertas")
def revisar_alertas(u: dict = Depends(get_usuario_actual)):
    """Verifica si falta el registro de hoy para mostrar alerta en frontend."""
    drogeria_id = u.get("drogeria_id")
    if not drogeria_id:
        return {"ok": True, "alerta": False}
    
    falta = verificar_alerta_condiciones(drogeria_id)
    return {"ok": True, "alerta": falta}


@router.get("/exportar")
def exportar_excel(
    mes: str = Query(..., description="Mes en formato YYYY-MM"),
    u: dict = Depends(get_usuario_actual)
):
    """Exporta los datos del mes a un archivo de Excel interactivo."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    drogeria_id = u.get("drogeria_id")
    if not drogeria_id:
        raise HTTPException(400, "Superadmin no puede exportar esto.")

    drog = get_drogeria(drogeria_id)
    nombre_drogueria = drog["nombre"] if drog else "Droguería"

    condiciones = obtener_condiciones_mes(drogeria_id, mes)
    
    # Crear Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Condiciones_{mes}"

    # Estilos
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Título superior
    ws.merge_cells("A1:G1")
    ws["A1"] = f"CONTROL DE TEMPERATURA Y HUMEDAD - {nombre_drogueria} ({mes})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align

    # Cabeceras
    headers = ["Día", "Temp AM (°C)", "Temp PM (°C)", "Hum AM (%)", "Hum PM (%)", "Firma AM", "Firma PM"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

    # Llenar datos (1 al 31)
    try:
        anio, mm = map(int, mes.split("-"))
        _, dias_mes = calendar.monthrange(anio, mm)
    except Exception:
        dias_mes = 31

    # Crear diccionario de acceso rápido
    datos_dict = {c["fecha"]: c for c in condiciones}

    fila = 4
    for dia in range(1, dias_mes + 1):
        fecha_str = f"{anio}-{mm:02d}-{dia:02d}"
        c = datos_dict.get(fecha_str, {})
        
        row_data = [
            dia,
            c.get("temperatura_am", ""),
            c.get("temperatura_pm", ""),
            c.get("humedad_am", ""),
            c.get("humedad_pm", ""),
            c.get("firma_am", ""),
            c.get("firma_pm", "")
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=fila, column=col_num, value=value)
            cell.alignment = center_align
            cell.border = thin_border
        fila += 1

    # Generar archivo en memoria
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"Control_Ambiental_{nombre_drogueria}_{mes}.xlsx"
    headers = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
